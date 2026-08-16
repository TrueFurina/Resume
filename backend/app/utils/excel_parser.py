"""解析实验安排表 Excel，构建排课占用数据"""

import openpyxl
from typing import List, Dict, Optional
from datetime import datetime

# 时间段映射（节次标记 → 实际时间）
TIME_SLOTS = {
    "567": "13:30-15:45",
    "678": "15:45-18:00",
    "91011": "18:30-20:45",
    "78": "15:50-18:05",
}

# 反查：实际时间 → 标准节次标记（Excel 里时间格式不统一，需归一化）
TIME_SLOT_NORMALIZE = {
    "13:30-15:45": "567",
    "15:45-18:00": "678",
    "18:30-20:45": "91011",
    "15:50-18:05": "78",
    "18:00-21:00": "91011",
    "14:00-17:00": "567",
    "9:00-11:15": "MORNING",
    "9:00-12:00": "MORNING",
    "567": "567",
    "678": "678",
    "91011": "91011",
    "78": "78",
}

WEEKDAY_MAP = {
    "周一": "monday", "周二": "tuesday", "周三": "wednesday",
    "周四": "thursday", "周五": "friday", "周六": "saturday", "周日": "sunday",
}

# 上午时段不算实验冲突（实验课都排在下午/晚上），筛选时忽略
IGNORED_TIME_SLOTS = {"MORNING", ""}


class ScheduleEntry:
    """一条排课记录"""
    def __init__(self, week: int, weekday: str, time_slot: str, class_name: str,
                 group: str, experiment: str, room: str, teacher: str):
        self.week = week
        self.weekday = weekday
        self.time_slot = time_slot
        self.class_name = class_name
        self.group = group
        self.experiment = experiment
        self.room = room
        self.teacher = teacher

    def key(self) -> str:
        """唯一标识一次排课占用的资源"""
        return f"{self.week}_{self.weekday}_{self.time_slot}_{self.room}"

    def __repr__(self):
        return f"[W{self.week} {self.weekday} {self.time_slot}] {self.class_name}/{self.group} @ {self.room}"


class ScheduleDatabase:
    """排课数据内存数据库"""

    def __init__(self):
        self.entries: List[ScheduleEntry] = []
        # 索引：week + weekday + time_slot → 该时段所有排课
        self._by_time: Dict[str, List[ScheduleEntry]] = {}
        # 索引：room → 该房间的所有排课
        self._by_room: Dict[str, List[ScheduleEntry]] = {}
        # 索引：teacher → 该教师的所有排课
        self._by_teacher: Dict[str, List[ScheduleEntry]] = {}
        # 索引：class_name → 该班级的所有排课
        self._by_class: Dict[str, List[ScheduleEntry]] = {}
        self.loaded = False

    def load_from_excel(self, filepath: str):
        """从 Excel 文件加载排课数据"""
        wb = openpyxl.load_workbook(filepath)
        self._filepath = filepath  # 保存路径，供 get_teacher_roles 等使用

        # 读取"大学物理实验安排表 (周) "工作表
        sheet = wb["大学物理实验安排表 (周) "]
        self.entries = []

        for row in sheet.iter_rows(min_row=3, values_only=True):
            week, weekday, time_slot, class_name, group, experiment, room, teacher = \
                row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]

            # 跳过空行和无效数据
            if not week or not isinstance(week, (int, float)):
                continue
            if not weekday or not class_name:
                continue

            # 清理教室字段
            room_str = str(room).strip() if room else ""
            teacher_str = str(teacher).strip() if teacher else ""
            time_raw = str(time_slot).strip() if time_slot else ""
            # 归一化时间格式：'13:30-15:45' → '567'，'、'等脏数据 → 跳过
            time_str = TIME_SLOT_NORMALIZE.get(time_raw, "")

            if not time_str:
                continue  # 无效或无法识别的时间段，跳过

            entry = ScheduleEntry(
                week=int(week),
                weekday=weekday.strip(),
                time_slot=time_str,
                class_name=str(class_name).strip(),
                group=str(group).strip() if group else "",
                experiment=str(experiment).strip() if experiment else "",
                room=room_str,
                teacher=teacher_str,
            )
            self.entries.append(entry)

        self._build_index()
        self.loaded = True
        print(f"✅ 已加载 {len(self.entries)} 条排课记录")

    def _build_index(self):
        """构建索引加速查询"""
        self._by_time.clear()
        self._by_room.clear()
        self._by_teacher.clear()
        self._by_class.clear()

        for e in self.entries:
            time_key = f"{e.week}_{e.weekday}_{e.time_slot}"
            self._by_time.setdefault(time_key, []).append(e)
            if e.room:
                self._by_room.setdefault(e.room, []).append(e)
            if e.teacher:
                self._by_teacher.setdefault(e.teacher, []).append(e)
            if e.class_name:
                self._by_class.setdefault(e.class_name, []).append(e)

    def get_occupied_slots(self, week: int) -> List[Dict]:
        """获取某周所有已占用的时段"""
        occupied = []
        for e in self.entries:
            if e.week == week:
                occupied.append({
                    "week": e.week,
                    "weekday": e.weekday,
                    "time_slot": e.time_slot,
                    "room": e.room,
                    "class_name": e.class_name,
                    "teacher": e.teacher,
                    "experiment": e.experiment,
                })
        return occupied

    def find_free_slots(self, week: int, weekdays: List[str],
                        time_slots: List[str], rooms: List[str]) -> List[Dict]:
        """找到空闲时间段"""
        occupied = set()
        for e in self.entries:
            if e.week == week:
                occupied.add(f"{e.weekday}_{e.time_slot}_{e.room}")

        free = []
        for wd in weekdays:
            for ts in time_slots:
                for rm in rooms:
                    if f"{wd}_{ts}_{rm}" not in occupied:
                        free.append({
                            "weekday": wd,
                            "time_slot": ts,
                            "room": rm,
                        })
        return free

    def get_all_rooms(self) -> List[str]:
        return sorted(set(e.room for e in self.entries if e.room))

    def get_all_teachers(self) -> List[str]:
        return sorted(set(e.teacher for e in self.entries if e.teacher))

    def get_teacher_roles(self) -> Dict[str, str]:
        """从"教师联系表"提取教师角色/备注（负责人、实验室管理员等）"""
        roles: Dict[str, str] = {}
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self._filepath if hasattr(self, "_filepath") else self.schedule_path())
            if "教师联系表" in wb.sheetnames:
                ws = wb["教师联系表"]
                for row in ws.iter_rows(min_row=3, values_only=True):
                    name, _, _, note = row[0], row[1], row[2], row[3]
                    if name and note:
                        roles[str(name).strip()] = str(note).strip()
        except Exception:
            pass
        return roles

    def schedule_path(self) -> str:
        """返回数据源文件路径（兼容）"""
        return getattr(self, "_filepath", "")

    def get_teacher_courses(self, teacher_name: str) -> List[Dict]:
        """
        按教师提取课程卡片（F2）
        将同一教师 + 同一课程名 + 同一班级的排课记录合并为一张课程卡片
        """
        # 时间段友好显示映射
        TIME_DISPLAY = {
            "567": "13:30-15:45", "678": "15:45-18:00",
            "91011": "18:30-20:45", "78": "15:50-18:05",
            "34": "10:00-11:40", "MORNING": "上午",
        }

        # 清理脏实验名（绪论课等提示性文本）
        def clean_experiment(name: str) -> str:
            if not name:
                return "未知实验"
            if "绪论" in name:
                return "绪论课"
            if "教室已安排" in name or "请查系统" in name:
                return "绪论课"
            return name

        # 按 (课程名, 班级) 分组
        groups: Dict[Tuple[str, str], Dict] = {}
        for e in self.entries:
            if not e.teacher or teacher_name not in e.teacher:
                continue
            key = (clean_experiment(e.experiment), e.class_name or "未知班级")
            if key not in groups:
                groups[key] = {
                    "course_name": key[0],
                    "class_name": key[1],
                    "slots": [],  # 所有上课时段
                    "weeks": set(),
                }
            groups[key]["slots"].append({
                "week": e.week,
                "weekday": e.weekday,
                "time_slot": e.time_slot,
                "room": e.room,
            })
            groups[key]["weeks"].add(e.week)

        # 转输出结构
        courses = []
        for info in groups.values():
            slots = info["slots"]
            # 排序取最早一周作为"原上课时间"代表
            slots_sorted = sorted(slots, key=lambda s: (s["week"], s["weekday"], s["time_slot"]))
            first = slots_sorted[0]
            weeks_sorted = sorted(info["weeks"])
            # 时间友好显示：周四 567 → 周四 13:30-15:45
            time_display = TIME_DISPLAY.get(first["time_slot"], first["time_slot"])
            courses.append({
                "course_name": info["course_name"],
                "class_name": info["class_name"],
                "original_time": f"{first['weekday']} {time_display}",
                "room": first["room"],
                "week_range": f"{weeks_sorted[0]}-{weeks_sorted[-1]}周",
                "total_slots": len(slots),
                # 可调状态：实验轮次固定课程默认可调（由前端/后续规则细化）
                "adjustable": True,
                "adjustable_reason": "",
            })
        return courses

    def get_all_classes(self) -> List[str]:
        return sorted(set(e.class_name for e in self.entries if e.class_name))

    def get_week_range(self) -> tuple:
        if not self.entries:
            return (0, 0)
        weeks = [e.week for e in self.entries]
        return (min(weeks), max(weeks))
